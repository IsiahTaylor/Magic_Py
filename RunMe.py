import pandas as pd
import requests
import time
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook

# Base URL for Scryfall API
SCRYFALL_API_URL = "https://api.scryfall.com/cards/named"

# File path for the Excel sheet
EXCEL_FILE = "mtg_collection.xlsx"

# Expected columns
EXPECTED_COLUMNS = ["Run", "Name", "Set", "Set #", "Quantity", "Price", "Total Price", "Last Updated"]

# Function to check if the file is available for writing
def wait_for_file(file_path, max_attempts=10, delay=3):
    """Wait until the file is available for writing"""
    attempts = 0
    while attempts < max_attempts:
        try:
            with open(file_path, "a"):  # Try opening in append mode
                return True
        except PermissionError:
            print(f"⚠️ File is locked! Waiting... ({attempts + 1}/{max_attempts})")
            time.sleep(delay)
            attempts += 1
    print("❌ File is still locked after multiple attempts. Exiting.")
    exit()

# Function to load the Excel sheet safely
def load_excel_sheet(file_path, sheet_name):
    """Loads an Excel sheet while ignoring the first row (TOTAL VALUE) and auto-correcting column count"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype={"Quantity": str})

        # Check if the first row contains "TOTAL VALUE" and ignore it during processing
        if str(df.iloc[0, 0]).strip().lower() == "total value":
            df = df.iloc[1:].reset_index(drop=True)  # Remove first row & reset index

        # If extra columns exist, automatically select only the first 8
        if len(df.columns) > len(EXPECTED_COLUMNS):
            print(f"⚠️ Warning: Found {len(df.columns)} columns but expected {len(EXPECTED_COLUMNS)}.")
            print("🔹 Automatically using the first 8 columns.")
            df = df.iloc[:, :len(EXPECTED_COLUMNS)]  # Keep only the first 8 columns

        # Check if column count still mismatches (if too few)
        if len(df.columns) != len(EXPECTED_COLUMNS):
            print(f"❌ Error: Expected {len(EXPECTED_COLUMNS)} columns but found {len(df.columns)} in '{sheet_name}'.")
            print("💡 Check if the sheet has missing columns.")
            print(f"📝 Found columns: {list(df.columns)}")
            exit()

        df.columns = EXPECTED_COLUMNS  # Rename columns dynamically
        return df

    except ValueError:
        print(f"❌ Error: The sheet name '{sheet_name}' was not found in '{file_path}'.")
        print("💡 Please check the sheet name and try again.")
        exit()

# Function to determine if a row should be updated based on the run mode
def should_update_row(row, run_mode):
    last_updated = pd.to_datetime(row["Last Updated"], errors="coerce")

    if run_mode == "all":
        return True
    elif run_mode == "checked" and row["Run"] == True:
        return True
    elif run_mode == "aged" and pd.notna(last_updated):
        return (datetime.today() - last_updated).days > 30
    elif run_mode == "empty":
        return pd.isna(row["Price"]) or row["Price"] == "Not Found"

    return False

# Function to perform an exact search for card price
def get_card_price(name, set_name=None, set_number=None):
    params = {"exact": name}
    if set_name:
        params["set"] = set_name.lower()
    if set_number:
        params["collector_number"] = set_number

    response = requests.get(SCRYFALL_API_URL, params=params)
    
    if response.status_code == 200:
        data = response.json()
        return float(data.get("prices", {}).get("usd", "N/A") or 0)

    print(f"⚠️ Exact search failed for '{name}', trying fuzzy search...")
    return fuzzy_search(name)

# Function to perform a fuzzy search if the exact match fails
def fuzzy_search(name):
    search_url = f"https://api.scryfall.com/cards/search?q={name.replace(' ', '+')}"
    response = requests.get(search_url)

    if response.status_code == 200:
        data = response.json()
        if "data" in data and data["data"]:
            best_match = data["data"][0]  # Take the first matching result
            return float(best_match.get("prices", {}).get("usd", "N/A") or 0)

    print(f"❌ No fuzzy match found for '{name}'.")
    return "Not Found"


# User inputs (Case-insensitive)
sheet_name = input("Enter the sheet name to use: ").strip()
run_mode = input("Enter run mode (all, checked, aged, empty): ").strip().lower()

# Validate run mode
valid_modes = {"all", "checked", "aged", "empty"}
if run_mode not in valid_modes:
    print(f"❌ Invalid run mode '{run_mode}'. Please enter one of: {', '.join(valid_modes)}")
    exit()


# Load the Excel file safely
df = load_excel_sheet(EXCEL_FILE, sheet_name)

print("✅ Sheet loaded successfully. Proceeding with updates...")

# Rename columns to match expected names
df.columns = ["Run", "Name", "Set", "Set #", "Quantity", "Price", "Total Price", "Last Updated"]

# Convert 'Run' column to boolean
df["Run"] = df["Run"].astype(bool)

# Convert 'Quantity' to integer (default to 1 if empty)
df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(1).astype(int)

# Convert 'Last Updated' column to datetime (Force MM/DD/YYYY format)
df["Last Updated"] = pd.to_datetime(df["Last Updated"], errors="coerce").dt.strftime("%m/%d/%Y")

# Remove old TOTAL VALUE row before adding a new one
df = df[df["Name"] != "TOTAL VALUE"]

# Select rows to update (Case-insensitive run mode)
rows_to_update = df[df.apply(lambda row: should_update_row(row, run_mode), axis=1)]

if not rows_to_update.empty:
    print(f"Updating {len(rows_to_update)} cards...")

# Process each card one by one
for index, row in rows_to_update.iterrows():
    card_name = row["Name"]
    set_name = row["Set"] if pd.notna(row["Set"]) and row["Set"].strip() else None
    set_number = row["Set #"] if pd.notna(row["Set #"]) and row["Set #"].strip() else None

    print(f"🔍 Searching for: {card_name} (Set: {set_name if set_name else 'Any'}, Set #: {set_number if set_number else 'Any'})")
    new_price = get_card_price(card_name, set_name, set_number)

    df.at[index, "Price"] = new_price
    df.at[index, "Total Price"] = new_price * row["Quantity"] if isinstance(new_price, float) else "Not Found"
    df.at[index, "Last Updated"] = datetime.today().strftime("%m/%d/%Y")  # Format as MM/DD/YYYY

    # Respect API rate limits
    time.sleep(0.1)

# Calculate the total collection value
total_sum = df["Total Price"].replace("Not Found", 0).sum()

# 🔹 Wait for the file to become available before writing 🔹
if not wait_for_file(EXCEL_FILE):
    exit()

# Load the workbook to **preserve formatting**
wb = load_workbook(EXCEL_FILE)
ws = wb[sheet_name]

# Overwrite the data in the sheet while preserving formatting
for row_idx, row in enumerate(df.itertuples(index=False), start=2):  # Start at row 2 (leave row 1 for total)
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Place TOTAL VALUE in **I1 & J1**
ws["I1"] = "TOTAL VALUE"
ws["J1"] = total_sum

# Save the workbook **without changing formatting**
wb.save(EXCEL_FILE)

print(f"✅ Updated {sheet_name} successfully. Total collection value: ${total_sum:.2f}")
